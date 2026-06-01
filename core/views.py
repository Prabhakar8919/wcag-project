from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Count, Q, Avg, Sum, F
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
import csv
import time
import json
import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse, urldefrag
from django.http import HttpResponse, JsonResponse

from .models import Project, Scan, Page
from rules.models import Issue
from crawler.tasks import crawl_and_analyze

# this starts the crawl task
@login_required
def crawler_view(request):
    if request.method == "POST":
        target_url = request.POST.get("url")
        wcag_level = request.POST.get("level", "AA")
        crawl_limit = int(request.POST.get("crawl_limit", 50))
        crawl_depth = int(request.POST.get("crawl_depth", 3))
        sitemap_enabled = request.POST.get("sitemap_enabled") == "on"
        estimated_pages = int(request.POST.get("estimated_pages", 0))
        
        user = request.user

        # Try to get existing project for this user, or create new one
        project, created = Project.objects.get_or_create(
            domain=target_url,
            user=user,
            defaults={
                'wcag_level': wcag_level,
                'crawl_limit': crawl_limit,
                'crawl_depth': crawl_depth,
                'sitemap_enabled': sitemap_enabled,
                'estimated_pages': estimated_pages
            }
        )
        
        # Update details if it changed for existing project
        if not created:
            project.wcag_level = wcag_level
            project.crawl_limit = crawl_limit
            project.crawl_depth = crawl_depth
            project.sitemap_enabled = sitemap_enabled
            project.estimated_pages = estimated_pages
            project.save()

        scan = Scan.objects.create(project=project, status='Pending')
        
        crawl_and_analyze.delay(scan.id)
        
        return redirect(f'/crawler/?project_id={project.id}')

    project_id = request.GET.get('project_id')
    context = {'project_id': project_id}
    if project_id:
        project = Project.objects.filter(id=project_id, user=request.user).first()
        if project:
            context['target_url'] = project.domain
            context['wcag_level'] = project.wcag_level
            
    return render(request, 'core/crawler.html', context)

# this shows global analysis across all projects
@login_required
def global_dashboard(request):
    total_pages = Page.objects.filter(scan__project__user=request.user).count()
    total_issues = Issue.objects.filter(scan__project__user=request.user).count()
    total_projects = Project.objects.filter(user=request.user).count()
    active_scans = Scan.objects.filter(project__user=request.user, status__in=['Pending', 'Crawling', 'Analyzing']).count()
    
    compliance_score = 100
    if total_pages > 0:
        penalty = min(total_issues / total_pages * 5, 100)
        compliance_score = int(100 - penalty)
        if compliance_score < 0: compliance_score = 0
    
    severity_counts = Issue.objects.filter(scan__project__user=request.user).values('severity').annotate(count=Count('severity'))
    severity_data = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0}
    for item in severity_counts:
        severity_data[item['severity']] = item['count']
        
    category_counts = list(Issue.objects.filter(scan__project__user=request.user).values('rule__category').annotate(count=Count('rule__category')).order_by('-count')[:10])
        
    recent_projects = Project.objects.filter(user=request.user).prefetch_related('scans', 'scans__report').order_by('-created_at')[:10]
    project_data = []
    
    trend_labels = []
    trend_data = []
    
    for p in recent_projects:
        # Avoid N+1 by using prefetched list
        scans = list(p.scans.all())
        latest_scan = scans[0] if scans else None
        status = latest_scan.status if latest_scan else "No scans yet"
        project_data.append({
            'project': p,
            'status': status
        })
        
    # Get last 5 for trends, utilizing prefetched Report to avoid N+1 queries
    for p in reversed(list(recent_projects)[:5]):
        scans = list(p.scans.all())
        latest_scan = scans[0] if scans else None
        if latest_scan and hasattr(latest_scan, 'report') and latest_scan.report:
            issue_count = latest_scan.report.total_issues_found
        else:
            issue_count = 0
        trend_labels.append(p.domain[:15] + '...')
        trend_data.append(issue_count)
        
    latest_project = Project.objects.filter(user=request.user).order_by('-created_at').first()
    latest_issue = Issue.objects.filter(scan__project__user=request.user).order_by('-created_at').first()
    total_ai_issues = Issue.objects.filter(scan__project__user=request.user, rule__check_type='llm').count()
        
    context = {
        'total_pages': total_pages,
        'total_issues': total_issues,
        'total_projects': total_projects,
        'active_scans': active_scans,
        'compliance_score': compliance_score,
        'severity_data': severity_data,
        'category_counts': category_counts,
        'projects': project_data[:5],
        'trend_labels': trend_labels,
        'trend_data': trend_data,
        'latest_issue': latest_issue,
        'total_ai_issues': total_ai_issues,
        'ai_avg_time': Scan.objects.filter(project__user=request.user, ai_pages_processed__gt=0).aggregate(avg=Avg(F('ai_total_time') / F('ai_pages_processed')))['avg'] or 0,
        'ai_success_rate': Scan.objects.filter(project__user=request.user).aggregate(total=Sum('ai_pages_processed'), errors=Sum('ai_errors_count'))
    }
    # Calculate success rate
    stats = context['ai_success_rate']
    if stats and stats['total'] and (stats['total'] + stats['errors']) > 0:
        context['ai_success_rate'] = (stats['total'] / (stats['total'] + stats['errors'])) * 100
    else:
        context['ai_success_rate'] = 100

    return render(request, 'core/global_dashboard.html', context)

import re

# formats raw LLaMA markdown outlines into styled, responsive HTML elements
def format_markdown_to_html(text):
    if not text:
        return ""
    # Convert headers (e.g. ### Header)
    text = re.sub(r'###\s*(.*?)(?:\n|$)', r'<h4 style="color: var(--accent-cyan); margin-top: 18px; margin-bottom: 10px; font-weight: 700; font-family: \'Space Grotesk\', sans-serif; font-size: 13px;">\1</h4>', text)
    text = re.sub(r'##\s*(.*?)(?:\n|$)', r'<h4 style="color: var(--accent-cyan); margin-top: 18px; margin-bottom: 10px; font-weight: 700; font-family: \'Space Grotesk\', sans-serif; font-size: 13px;">\1</h4>', text)
    text = re.sub(r'#\s*(.*?)(?:\n|$)', r'<h4 style="color: var(--accent-cyan); margin-top: 18px; margin-bottom: 10px; font-weight: 700; font-family: \'Space Grotesk\', sans-serif; font-size: 13px;">\1</h4>', text)
    
    # Convert bold text (e.g. **bold**)
    text = re.sub(r'\*\*(.*?)\*\*', r'<strong style="color: #FFFFFF; font-weight: 600;">\1</strong>', text)
    
    # Convert list items (e.g. - item)
    text = re.sub(r'^\s*[-*•]\s*(.*?)$', r'<div style="margin-left: 12px; margin-bottom: 6px; display: flex; align-items: flex-start; gap: 8px; color: var(--text-secondary); font-size: 12.5px;"><span style="color: var(--accent-cyan); flex-shrink: 0;">•</span><span>\1</span></div>', text, flags=re.MULTILINE)
    return text

# this shows full analysis results
@login_required
def dashboard_view(request, project_id):
    project = get_object_or_404(Project, id=project_id, user=request.user)
    latest_scan = project.scans.prefetch_related('pages').first()
    
    report = None
    if latest_scan:
        pages = latest_scan.pages.all()
        total_pages = pages.count()
        
        # Get the compiled Report if available
        if hasattr(latest_scan, 'report'):
            report = latest_scan.report
            total_issues = report.total_issues_found
            compliance_score = int(report.score)
        else:
            total_issues = Issue.objects.filter(scan=latest_scan).count()
            compliance_score = 100
            if total_pages > 0:
                penalty = min(total_issues / total_pages * 5, 100)
                compliance_score = int(100 - penalty)
                if compliance_score < 0: compliance_score = 0
                
        severity_counts = Issue.objects.filter(scan=latest_scan).values('severity').annotate(count=Count('severity'))
        severity_data = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0}
        for item in severity_counts:
            severity_data[item['severity']] = item['count']
            
        category_counts = list(Issue.objects.filter(scan=latest_scan).values('rule__category').annotate(count=Count('rule__category')).order_by('-count')[:5])
        ai_issues = Issue.objects.filter(scan=latest_scan, rule__check_type='llm').select_related('rule', 'page')
            
    else:
        pages = []
        total_issues = 0
        severity_data = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0}
        category_counts = []
        ai_issues = []
        compliance_score = 100

    ai_summary_html = ""
    ai_health_html = ""
    ai_legal_html = ""
    ai_risk_html = ""
    
    if latest_scan and report:
        ai_summary_html = format_markdown_to_html(report.ai_summary)
        ai_health_html = format_markdown_to_html(report.ai_health_report)
        ai_legal_html = format_markdown_to_html(report.ai_legal_insights)
        ai_risk_html = format_markdown_to_html(report.ai_risk_analysis)

    context = {
        'project': project,
        'scan': latest_scan,
        'report': report,
        'pages': pages,
        'total_issues': total_issues,
        'severity_data': severity_data,
        'category_counts': category_counts,
        'compliance_score': compliance_score,
        'ai_issues': ai_issues,
        'ai_pages_processed': latest_scan.ai_pages_processed if latest_scan else 0,
        'ai_avg_time': (latest_scan.ai_total_time / latest_scan.ai_pages_processed) if latest_scan and latest_scan.ai_pages_processed > 0 else 0,
        'ai_errors': latest_scan.ai_errors_count if latest_scan else 0,
        'ai_summary_html': ai_summary_html,
        'ai_health_html': ai_health_html,
        'ai_legal_html': ai_legal_html,
        'ai_risk_html': ai_risk_html,
    }
    return render(request, 'core/dashboard.html', context)

# this lists all scans
@login_required
def projects_list(request):
    # Optimize with prefetch_related for scans and their reports
    projects = Project.objects.filter(user=request.user).prefetch_related('scans', 'scans__report').order_by('-created_at')
    
    project_data = []
    for p in projects:
        scans = list(p.scans.all())
        latest_scan = scans[0] if scans else None
        status = latest_scan.status if latest_scan else "No scans yet"
        
        # calculate compliance for project using Report if available to save DB hits
        if latest_scan and hasattr(latest_scan, 'report'):
            pages_count = latest_scan.report.total_pages_scanned
            issues_count = latest_scan.report.total_issues_found
        else:
            pages_count = latest_scan.pages.count() if latest_scan else 0
            issues_count = Issue.objects.filter(scan=latest_scan).count() if latest_scan else 0
        
        compliance_score = 100
        if pages_count > 0:
            penalty = min(issues_count / pages_count * 5, 100)
            compliance_score = int(100 - penalty)
            if compliance_score < 0: compliance_score = 0
            
        project_data.append({
            'project': p,
            'status': status,
            'compliance_score': compliance_score
        })

    return render(request, 'core/projects.html', {'projects': project_data})

from django.views.decorators.http import require_POST

# this deletes a project
@login_required
@require_POST
def delete_project(request, project_id):
    project = get_object_or_404(Project, id=project_id, user=request.user)
    project.delete()
    return redirect('projects')


def get_registered_domain(netloc):
    netloc = netloc.replace('www.', '').lower()
    parts = netloc.split('.')
    if len(parts) >= 2:
        # Check for co.uk, com.au, org.in, etc.
        if len(parts[-2]) <= 3 and parts[-1] in ('uk', 'jp', 'au', 'nz', 'za', 'br', 'mx', 'in', 'cn', 'org'):
            if len(parts) >= 3:
                return '.'.join(parts[-3:])
        return '.'.join(parts[-2:])
    return netloc

def fetch_html_with_fallback(url, headers):
    user_agent = headers.get('User-Agent', 'Mozilla/5.0')
    import subprocess
    try:
        cmd = ["curl.exe", "-s", "-L", "-A", user_agent, url]
        res = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8', errors='ignore', timeout=5)
        if res.returncode == 0 and res.stdout:
            return res.stdout
    except Exception:
        pass
    return None

# this estimates how many pages exist
def estimate_pages_api(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            target_url = data.get('url')
        except:
            target_url = request.POST.get('url')

        if not target_url:
            return JsonResponse({'error': 'URL required'}, status=400)

        if not target_url.startswith(('http://', 'https://')):
            target_url = 'https://' + target_url

        try:
            parsed_domain = urlparse(target_url)
            base_domain = get_registered_domain(parsed_domain.netloc)
            
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
            
            # Try to fetch and count from sitemap.xml first for exact count
            sitemap_urls = set()
            sitemap_loc = urljoin(target_url, "/sitemap.xml")
            sitemap_text = None
            try:
                sitemap_response = requests.get(sitemap_loc, headers=headers, timeout=3, verify=False)
                if sitemap_response.status_code == 200:
                    sitemap_text = sitemap_response.text
                else:
                    sitemap_text = fetch_html_with_fallback(sitemap_loc, headers)
            except Exception:
                sitemap_text = fetch_html_with_fallback(sitemap_loc, headers)

            if sitemap_text:
                locs = re.findall(r'<loc>\s*(https?://[^\s<]+)\s*</loc>', sitemap_text, re.IGNORECASE)
                
                sub_sitemap_urls = []
                for loc in locs:
                    if 'sitemap' in loc.lower() and loc.endswith('.xml'):
                        sub_sitemap_urls.append(loc)
                    else:
                        sitemap_urls.add(loc)
                
                # Fetch sub-sitemaps concurrently using ThreadPoolExecutor, limited to top 3 sub-sitemaps
                if sub_sitemap_urls:
                    sub_sitemap_urls = sub_sitemap_urls[:3]
                    def fetch_sub_sitemap(url):
                        try:
                            res = requests.get(url, headers=headers, timeout=1.5, verify=False)
                            if res.status_code == 200:
                                return res.text
                        except:
                            pass
                        return None
                    
                    import concurrent.futures
                    with concurrent.futures.ThreadPoolExecutor(max_workers=3) as sub_executor:
                        results = sub_executor.map(fetch_sub_sitemap, sub_sitemap_urls)
                        for sub_text in results:
                            if sub_text:
                                sub_locs = re.findall(r'<loc>\s*(https?://[^\s<]+)\s*</loc>', sub_text, re.IGNORECASE)
                                sitemap_urls.update(sub_locs)

            # Filter sitemap URLs to match the target base domain, capped at 150 for performance
            filtered_sitemap_urls = set()
            for u in sitemap_urls:
                if len(filtered_sitemap_urls) >= 150:
                    break
                u_no_frag, _ = urldefrag(u)
                u_netloc = get_registered_domain(urlparse(u_no_frag).netloc)
                if u_netloc == base_domain:
                    filtered_sitemap_urls.add(u_no_frag)

            if filtered_sitemap_urls:
                # Returns the exact page count from the sitemap!
                return JsonResponse({'estimated_pages': len(filtered_sitemap_urls)})

            # Fallback to standard crawling estimator, parallelized with ThreadPoolExecutor
            import concurrent.futures
            import urllib3
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

            visited = {target_url}
            pages_checked = 0
            
            start_time = time.time()
            max_duration = 3.0  # Snappy timeout limit
            max_pages_to_check = 80 # Higher scan limit for accuracy

            def fetch_links(url):
                try:
                    response = requests.get(url, headers=headers, timeout=1.5, verify=False)
                    if response.status_code == 200:
                        soup = BeautifulSoup(response.text, 'html.parser')
                        found = []
                        for a_tag in soup.find_all('a', href=True):
                            link = a_tag['href'].strip()
                            if not link or link.startswith(('javascript:', 'mailto:', 'tel:', '#')):
                                continue
                            abs_url = urljoin(response.url, link)
                            url_no_fragment, _ = urldefrag(abs_url)
                            netloc = get_registered_domain(urlparse(url_no_fragment).netloc)
                            if netloc == base_domain:
                                found.append(url_no_fragment)
                        return found
                except Exception:
                    pass
                return []

            with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
                futures = {executor.submit(fetch_links, target_url): target_url}
                
                while futures and (time.time() - start_time) < max_duration and len(visited) < max_pages_to_check:
                    done, _ = concurrent.futures.wait(futures, return_when=concurrent.futures.FIRST_COMPLETED, timeout=0.1)
                    for fut in done:
                        url = futures.pop(fut)
                        pages_checked += 1
                        try:
                            links = fut.result()
                            for link in links:
                                if link not in visited and len(visited) < max_pages_to_check:
                                    visited.add(link)
                                    futures[executor.submit(fetch_links, link)] = link
                        except Exception:
                            pass

            estimated_count = len(visited)
            if estimated_count == 0 and pages_checked > 0:
                estimated_count = 1

            return JsonResponse({'estimated_pages': estimated_count})

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
            
    return JsonResponse({'error': 'Invalid request'}, status=400)

# this updates live crawling data
@login_required
def crawl_status_api(request, project_id):
    project = get_object_or_404(Project, id=project_id, user=request.user)
    latest_scan = project.scans.first()
    
    if not latest_scan:
        return JsonResponse({"status": "Not Found"})
        
    pages_crawled = Page.objects.filter(scan=latest_scan, status='Crawled').count()
    issues_count = Issue.objects.filter(scan=latest_scan).count()
    
    # Optmize: only fetch last 5 URLs, flatten to strings directly
    current_urls = list(Page.objects.filter(scan=latest_scan).order_by('-created_at')[:5].values_list('url', flat=True))
    
    return JsonResponse({
        "status": latest_scan.status,
        "pages_crawled": pages_crawled,
        "current_urls": current_urls,
        "issues_count": issues_count,
        "crawl_limit": project.crawl_limit,
        "estimated_pages": project.estimated_pages,
        "ai_pages_processed": latest_scan.ai_pages_processed,
        "ai_errors_count": latest_scan.ai_errors_count
    })

# this shows issues for a specific page
@login_required
def page_detail(request, page_id):
    page = get_object_or_404(Page, id=page_id, scan__project__user=request.user)
    issues = page.issues.all()
    has_fixes = any(issue.corrected_html for issue in issues)
    
    context = {
        'page': page,
        'issues': issues,
        'has_fixes': has_fixes,
    }
    return render(request, 'core/page.html', context)

# this exports data to CSV
@login_required
def export_csv(request, project_id):
    project = get_object_or_404(Project, id=project_id, user=request.user)
    latest_scan = project.scans.first()
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="issues_{project.domain.replace("https://", "").replace("http://", "").replace("/", "_")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Page URL', 'WCAG Rule ID', 'Issue Message', 'Severity', 'Fix Suggestion'])
    
    if latest_scan:
        issues = Issue.objects.filter(scan=latest_scan).select_related('page', 'rule')
        for issue in issues:
            writer.writerow([
                issue.page.url,
                issue.rule.wcag_id if issue.rule else '',
                issue.message,
                issue.severity.capitalize(),
                issue.fix_suggestion or (issue.rule.fix_suggestion if issue.rule else '')
            ])
            
    return response


# ==================================================
# AUTHENTICATION & USER MANAGEMENT VIEWS
# ==================================================
from django.contrib.auth import login as auth_login, logout as auth_logout, authenticate
from django.contrib.auth.models import User
from django.contrib import messages
from django.utils import timezone
from .models import Profile
from .otp_utils import setup_user_otp, send_otp_email, verify_otp_code, send_otp_email_async

def signup_view(request):
    if request.user.is_authenticated:
        return redirect('home')
        
    if request.method == 'POST':
        full_name = request.POST.get('full_name', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        
        if not full_name or not email or not password:
            return render(request, 'core/auth/signup.html', {'error': 'All fields are required.'})
            
        if password != confirm_password:
            return render(request, 'core/auth/signup.html', {'error': 'Passwords do not match.'})
            
        if User.objects.filter(username=email).exists() or User.objects.filter(email=email).exists():
            return render(request, 'core/auth/signup.html', {'error': 'A user with this email already exists.'})
            
        try:
            # Create user safely
            first_name = full_name.split(' ', 1)[0]
            last_name = full_name.split(' ', 1)[1] if ' ' in full_name else ''
            
            user = User.objects.create_user(
                username=email,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name
            )
            
            # Create profile
            profile, _ = Profile.objects.get_or_create(user=user)
            
            # Setup user OTP and send email
            raw_otp = setup_user_otp(profile)
            send_otp_email_async(user, raw_otp)
            
            # Store target email in session
            request.session['pre_verified_user_email'] = email
            request.session['otp_action'] = 'verify'
            
            messages.success(request, "Account created successfully! We have sent a 6-digit verification code to your email.")
            return redirect('verify_otp')
        except Exception as e:
            return render(request, 'core/auth/signup.html', {'error': f'Failed to create account: {str(e)}'})
            
    return render(request, 'core/auth/signup.html')

def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')
        
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password')
        remember_me = request.POST.get('remember_me')
        
        if not email or not password:
            return render(request, 'core/auth/login.html', {'error': 'Email and password are required.'})
            
        user = authenticate(request, username=email, password=password)
        if user is None:
            try:
                user_by_email = User.objects.get(email=email)
            except User.DoesNotExist:
                user_by_email = None
            if user_by_email:
                user = authenticate(request, username=user_by_email.username, password=password)

        if user is not None:
            profile, _ = Profile.objects.get_or_create(user=user)
            
            # Reject login if email is not verified
            raw_otp = setup_user_otp(profile)
            send_otp_email_async(user, raw_otp)

            request.session['pre_verified_user_email'] = user.email
            request.session['otp_action'] = 'login'
            request.session['remember_me'] = bool(remember_me)
            messages.info(request, "A verification code has been sent to your email. Enter it to complete login.")
            return redirect('verify_otp')
        else:
            return render(request, 'core/auth/login.html', {'error': 'Invalid email or password.'})
            
    return render(request, 'core/auth/login.html')

def logout_view(request):
    auth_logout(request)
    return redirect('login')

def verify_otp_view(request):
    if request.user.is_authenticated and hasattr(request.user, 'profile') and request.user.profile.is_email_verified:
        return redirect('home')
        
    email = request.session.get('pre_verified_user_email')
    if not email:
        messages.error(request, "No verification session found. Please log in or sign up first.")
        return redirect('login')
        
    user = get_object_or_404(User, email=email)
    profile, _ = Profile.objects.get_or_create(user=user)
    
    if request.method == 'POST':
        otp_digits = [
            request.POST.get('otp_1', ''),
            request.POST.get('otp_2', ''),
            request.POST.get('otp_3', ''),
            request.POST.get('otp_4', ''),
            request.POST.get('otp_5', ''),
            request.POST.get('otp_6', ''),
        ]
        raw_otp = "".join(otp_digits).strip()
        
        if not raw_otp:
            raw_otp = request.POST.get('otp', '').strip()
            
        if len(raw_otp) != 6 or not raw_otp.isdigit():
            return render(request, 'core/auth/verify_otp.html', {
                'email': email,
                'error': 'Please enter a valid 6-digit verification code.'
            })
            
        otp_action = request.session.get('otp_action', 'login' if profile.is_email_verified else 'verify')
        require_existing = otp_action == 'login'
        success, message = verify_otp_code(profile, raw_otp, require_existing_verification=require_existing)
        
        if success:
            auth_login(request, user)
            remember_me = request.session.pop('remember_me', False)
            if remember_me:
                request.session.set_expiry(1209600)  # 2 weeks
            else:
                request.session.set_expiry(0)  # session expires on browser close
            request.session.pop('pre_verified_user_email', None)
            request.session.pop('otp_action', None)
            if otp_action == 'verify':
                messages.success(request, "Your email has been verified successfully!")
            else:
                messages.success(request, "Login successful. You are now signed in.")
            return redirect('home')
        else:
            is_locked = "locked" in message.lower() or "too many" in message.lower()
            return render(request, 'core/auth/verify_otp.html', {
                'email': email,
                'error': message,
                'is_locked': is_locked
            })
            
    remaining_seconds = 600
    if profile.otp_expiry:
        delta = profile.otp_expiry - timezone.now()
        remaining_seconds = max(0, int(delta.total_seconds()))
        
    return render(request, 'core/auth/verify_otp.html', {
        'email': email,
        'remaining_seconds': remaining_seconds,
    })

@require_POST
def resend_otp_view(request):
    email = request.session.get('pre_verified_user_email')
    if not email:
        return JsonResponse({'success': False, 'message': 'No active verification session found.'}, status=400)
        
    user = User.objects.filter(email=email).first()
    if not user:
        return JsonResponse({'success': False, 'message': 'User not found.'}, status=404)
        
    profile, _ = Profile.objects.get_or_create(user=user)
    
    # 60s resend cooldown enforcement
    if profile.otp_last_resent:
        elapsed = timezone.now() - profile.otp_last_resent
        if elapsed.total_seconds() < 60:
            remaining = 60 - int(elapsed.total_seconds())
            return JsonResponse({
                'success': False,
                'message': f'Please wait {remaining} seconds before requesting a new code.'
            }, status=429)
            
    raw_otp = setup_user_otp(profile)
    profile.otp_last_resent = timezone.now()
    profile.save()
    
    send_otp_email_async(user, raw_otp)
    
    return JsonResponse({
        'success': True,
        'message': 'A new verification code has been successfully sent to your email.'
    })



# ==================================================
# WAVE-LIKE VISUAL ACCESSIBILITY OVERLAY
# ==================================================
from django.conf import settings
from django.views.decorators.clickjacking import xframe_options_sameorigin

@xframe_options_sameorigin
@login_required
def page_overlay_view(request, page_id):
    page = get_object_or_404(Page, id=page_id, scan__project__user=request.user)
    html = page.html_snapshot or "<html><body><h3>No snapshot captured for this page.</h3></body></html>"
    
    # Parse base domain
    parsed = urlparse(page.url)
    base_url = f"{parsed.scheme}://{parsed.netloc}"
    
    # Assemble issue details for the script engine
    issues = []
    for issue in page.issues.all():
        issues.append({
            'rule_title': issue.rule.title if issue.rule else 'AI Usability Issue',
            'wcag_id': issue.rule.wcag_id if issue.rule else 'LLM_UX',
            'rule_id': issue.rule.wcag_id if issue.rule else 'LLM_UX',
            'severity': issue.severity,
            'message': issue.message,
            'element_html': issue.element_html,
            'fix_suggestion': issue.fix_suggestion or (issue.rule.fix_suggestion if issue.rule else ''),
            'corrected_html': issue.corrected_html
        })
        
    issues_json = json.dumps(issues)
    
    # Parse and modify html snapshot
    soup = BeautifulSoup(html, 'html.parser')
    
    # Inject base href tag
    base_tag = soup.new_tag('base', href=base_url)
    if soup.head:
        soup.head.insert(0, base_tag)
    elif soup.html:
        head = soup.new_tag('head')
        head.append(base_tag)
        soup.html.insert(0, head)
        
    # Inject overlay client engine
    overlay_script = soup.new_tag('script', src=f"{settings.STATIC_URL}core/js/overlay_engine.js")
    if soup.body:
        soup.body.append(overlay_script)
    else:
        soup.append(overlay_script)
        
    # Trigger overlay rendering
    trigger_code = f"""
    document.addEventListener("DOMContentLoaded", function() {{
        if (window.renderWCAGOverlays) {{
            window.renderWCAGOverlays({issues_json});
        }} else {{
            setTimeout(function() {{
                if (window.renderWCAGOverlays) window.renderWCAGOverlays({issues_json});
            }}, 500);
        }}
    }});
    """
    trigger_script = soup.new_tag('script')
    trigger_script.string = trigger_code
    if soup.body:
        soup.body.append(trigger_script)
    else:
        soup.append(trigger_script)
        
    return HttpResponse(str(soup), content_type="text/html")


# ==================================================
# ENTERPRISE EXPORT ENGINE (PDF & EXCEL REPORTS)
# ==================================================
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

@login_required
def export_pdf(request, project_id):
    project = get_object_or_404(Project, id=project_id, user=request.user)
    latest_scan = project.scans.first()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="accessibility_report_{project.domain.replace("https://", "").replace("http://", "").replace("/", "_")}.pdf"'
    
    # Initialize Doc
    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    story = []
    styles = getSampleStyleSheet()
    
    # Custom Corporate Styles
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor('#0F172A'),
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        'SubtitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#64748B'),
        spaceAfter=20
    )
    h2_style = ParagraphStyle(
        'Heading2Style',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=colors.HexColor('#1E293B'),
        spaceBefore=15,
        spaceAfter=10
    )
    body_style = ParagraphStyle(
        'BodyStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9.5,
        textColor=colors.HexColor('#334155'),
        spaceAfter=8,
        leading=14
    )
    
    # Report Header
    story.append(Paragraph("WCAG Accessibility Compliance Report", title_style))
    story.append(Paragraph(f"Target Domain: {project.domain}  |  Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S UTC')}", subtitle_style))
    story.append(Spacer(1, 10))
    
    report = getattr(latest_scan, 'report', None)
    score = int(report.score) if report else 100
    
    # Summary Table
    summary_data = [
        [Paragraph("<b>Audit Metric</b>", body_style), Paragraph("<b>Score / Value</b>", body_style), Paragraph("<b>Status</b>", body_style)],
        [Paragraph("Overall Accessibility Score", body_style), Paragraph(f"<b>{score} / 100</b>", body_style), Paragraph("Excellent" if score > 90 else ("Pass" if score > 75 else "Needs Improvement"), body_style)],
        [Paragraph("Total Pages Crawled", body_style), Paragraph(str(report.total_pages_scanned if report else latest_scan.pages.count()), body_style), Paragraph("N/A", body_style)],
        [Paragraph("Total Violations Detected", body_style), Paragraph(str(report.total_issues_found if report else Issue.objects.filter(scan=latest_scan).count()), body_style), Paragraph("Action Required" if (report.total_issues_found if report else 1) > 0 else "Compliant", body_style)]
    ]
    
    t_summary = Table(summary_data, colWidths=[200, 150, 150])
    t_summary.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F8FAFC')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(t_summary)
    story.append(Spacer(1, 15))
    
    if report:
        # POUR Scoreboard
        story.append(Paragraph("POUR Principle Compliance Checklist", h2_style))
        pour_data = [
            [Paragraph("<b>POUR Category</b>", body_style), Paragraph("<b>Grade</b>", body_style)],
            [Paragraph("<b>👁️ Perceivable</b> (Images, headings, alt attributes)", body_style), Paragraph(f"{int(report.score_perceivable)} / 100", body_style)],
            [Paragraph("<b>⌨️ Operable</b> (Keyboard access, focus markers, page layouts)", body_style), Paragraph(f"{int(report.score_operable)} / 100", body_style)],
            [Paragraph("<b>🧠 Understandable</b> (Input labels, error handling, clean structures)", body_style), Paragraph(f"{int(report.score_understandable)} / 100", body_style)],
            [Paragraph("<b>🛠️ Robust</b> (Standard parsing, ARIA attributes, semantic integrity)", body_style), Paragraph(f"{int(report.score_robust)} / 100", body_style)],
        ]
        t_pour = Table(pour_data, colWidths=[350, 150])
        t_pour.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F8FAFC')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(t_pour)
        story.append(Spacer(1, 15))

    # Violations Details
    story.append(Paragraph("Detailed Discovered Violations Log", h2_style))
    if latest_scan:
        issues = Issue.objects.filter(scan=latest_scan).select_related('page', 'rule')[:50]  # limit to top 50 in pdf
        
        issue_rows = [
            [Paragraph("<b>WCAG Rule</b>", body_style), Paragraph("<b>Severity</b>", body_style), Paragraph("<b>Page URL</b>", body_style), Paragraph("<b>Remediation roadmap</b>", body_style)]
        ]
        
        for issue in issues:
            rule_id = issue.rule.wcag_id if issue.rule else 'LLM'
            issue_rows.append([
                Paragraph(f"Rule {rule_id}<br/><font color='#64748b'>{issue.rule.category if issue.rule else 'AI Semantics'}</font>", body_style),
                Paragraph(f"<font color='{'#e11d48' if issue.severity == 'critical' else ('#d97706' if issue.severity == 'high' else '#2563eb')}'><b>{issue.severity.upper()}</b></font>", body_style),
                Paragraph(issue.page.url.replace("https://", "").replace("http://", "")[:25] + "...", body_style),
                Paragraph(f"{issue.message}<br/><b>Fix:</b> {issue.fix_suggestion or issue.rule.fix_suggestion}", body_style)
            ])
            
        t_issues = Table(issue_rows, colWidths=[90, 70, 110, 230])
        t_issues.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F8FAFC')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(t_issues)
    else:
        story.append(Paragraph("No scan issues logged.", body_style))
        
    doc.build(story)
    return response

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

@login_required
def export_excel(request, project_id):
    project = get_object_or_404(Project, id=project_id, user=request.user)
    latest_scan = project.scans.first()
    report = getattr(latest_scan, 'report', None)
    
    wb = Workbook()
    
    # Custom Fonts and Fills
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    title_font = Font(name='Segoe UI', size=16, bold=True, color='0F172A')
    section_font = Font(name='Segoe UI', size=12, bold=True, color='1E293B')
    bold_font = Font(name='Segoe UI', size=10, bold=True)
    normal_font = Font(name='Segoe UI', size=10)
    
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    ai_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # Sheet 1: Executive Audit Summary
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1['A1'] = "WCAG SaaS Audit Executive Report"
    ws1['A1'].font = title_font
    
    ws1['A3'] = "Metadata"
    ws1['A3'].font = section_font
    ws1.append(["Domain Name", project.domain])
    ws1.append(["Scan Date", timezone.now().strftime('%Y-%m-%d %H:%M UTC')])
    ws1.append(["WCAG Target Level", f"WCAG {project.wcag_level}"])
    ws1.append(["Total Pages Scanned", report.total_pages_scanned if report else latest_scan.pages.count()])
    ws1.append(["Total Accessibility Issues", report.total_issues_found if report else 0])
    ws1.append(["Overall Accessibility Score", f"{int(report.score if report else 100)} / 100"])
    
    for r in range(4, 10):
        ws1.cell(row=r, column=1).font = bold_font
        ws1.cell(row=r, column=2).font = normal_font
        ws1.cell(row=r, column=1).border = thin_border
        ws1.cell(row=r, column=2).border = thin_border
        
    ws1['A11'] = "POUR Principles Scorecards"
    ws1['A11'].font = section_font
    
    ws1.append([]) # empty
    ws1.append(["Principle", "Description", "Compliance Score"])
    
    pour_rows = [
        ["Perceivable", "Alt texts, semantic hierarchies, markup structures", f"{int(report.score_perceivable if report else 100)}%"],
        ["Operable", "Keyboard navigation, focus visible tags, bypass anchors", f"{int(report.score_operable if report else 100)}%"],
        ["Understandable", "Language identifiers, input tags, labels, instructions", f"{int(report.score_understandable if report else 100)}%"],
        ["Robust", "HTML validation tags, custom attributes, ARIA tags", f"{int(report.score_robust if report else 100)}%"],
    ]
    
    for row in pour_rows:
        ws1.append(row)
        
    # Format table headers
    for c in range(1, 4):
        cell = ws1.cell(row=13, column=c)
        cell.font = header_font
        cell.fill = header_fill
        
    for r in range(14, 18):
        ws1.cell(row=r, column=1).font = bold_font
        ws1.cell(row=r, column=2).font = normal_font
        ws1.cell(row=r, column=3).font = bold_font
        for c in range(1, 4):
            ws1.cell(row=r, column=c).border = thin_border

    # Sheet 2: Audited Pages Checklist
    ws2 = wb.create_sheet(title="Audited Pages")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.append(["URL", "Page Title", "HTTP Response Code", "Issue Counts", "Compliance Status"])
    for page in latest_scan.pages.all():
        count = page.issues.count()
        status = "Pass" if count == 0 else ("Warning" if count < 5 else "Fail")
        ws2.append([page.url, page.title or 'No Title', page.status_code or 200, count, status])
        
    for col in range(1, 6):
        cell = ws2.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        
    for r in range(2, ws2.max_row + 1):
        for c in range(1, 6):
            cell = ws2.cell(row=r, column=c)
            cell.font = normal_font
            cell.border = thin_border
            if r % 2 == 0:
                cell.fill = zebra_fill

    # Sheet 3: Discovered Violations Log
    ws3 = wb.create_sheet(title="Violations Log")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.append(["Page URL", "Rule ID", "Rule Title", "Category", "Level", "Severity", "Violation Message", "Remediation Recommendation", "AI Auto-Fix Code"])
    
    issues = Issue.objects.filter(scan=latest_scan).select_related('page', 'rule')
    for issue in issues:
        rule_id = issue.rule.wcag_id if issue.rule else 'LLM'
        rule_title = issue.rule.title if issue.rule else 'AI Usability'
        category = issue.rule.category if issue.rule else 'AI Insights'
        level = issue.rule.level if issue.rule else 'AA'
        
        ws3.append([
            issue.page.url,
            rule_id,
            rule_title,
            category,
            level,
            issue.severity.upper(),
            issue.message,
            issue.fix_suggestion or issue.rule.fix_suggestion if issue.rule else '',
            issue.corrected_html or ''
        ])
        
    for col in range(1, 10):
        cell = ws3.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        
    for r in range(2, ws3.max_row + 1):
        for c in range(1, 10):
            cell = ws3.cell(row=r, column=c)
            cell.font = normal_font
            cell.border = thin_border
            if r % 2 == 0:
                cell.fill = zebra_fill
            if c == 9 and cell.value:
                cell.fill = ai_fill
                
    # Auto-fit column widths across sheets
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)
            
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="accessibility_audit_{project.domain.replace("https://", "").replace("http://", "").replace("/", "_")}.xlsx"'
    wb.save(response)
    return response


@login_required
def get_neural_issues_api(request, project_id):
    project = get_object_or_404(Project, id=project_id, user=request.user)
    latest_scan = project.scans.first()
    if not latest_scan:
        return JsonResponse({"issues": [], "ai_pages_processed": 0, "total_pages": 0, "ai_complete": True})
        
    ai_issues = Issue.objects.filter(scan=latest_scan, rule__check_type='llm').select_related('rule', 'page').order_by('created_at')
    
    issues_list = []
    for issue in ai_issues:
        issues_list.append({
            'id': issue.id,
            'severity': issue.severity,
            'rule_title': issue.rule.title if issue.rule else 'Usability Friction',
            'wcag_id': issue.rule.wcag_id if issue.rule else 'LLM_UX',
            'message': issue.message,
            'page_url': issue.page.url,
            'page_detail_url': f"/page/{issue.page.id}/",
            'fix_suggestion': issue.fix_suggestion,
            'corrected_html': issue.corrected_html
        })
        
    total_pages = latest_scan.pages.count() if latest_scan else 0
    ai_pages = latest_scan.ai_pages_processed if latest_scan else 0
    ai_errors = latest_scan.ai_errors_count if latest_scan else 0
    
    # Also fetch formatted summaries if the scan is fully complete
    ai_complete = (ai_pages + ai_errors) >= total_pages if latest_scan else True
    ai_summary_html = ""
    ai_health_html = ""
    ai_legal_html = ""
    ai_risk_html = ""
    
    if ai_complete and latest_scan and hasattr(latest_scan, 'report'):
        report = latest_scan.report
        ai_summary_html = format_markdown_to_html(report.ai_summary)
        ai_health_html = format_markdown_to_html(report.ai_health_report)
        ai_legal_html = format_markdown_to_html(report.ai_legal_insights)
        ai_risk_html = format_markdown_to_html(report.ai_risk_analysis)
        
    return JsonResponse({
        "issues": issues_list,
        "ai_pages_processed": ai_pages,
        "ai_errors_count": ai_errors,
        "total_pages": total_pages,
        "ai_complete": ai_complete,
        "ai_summary_html": ai_summary_html,
        "ai_health_html": ai_health_html,
        "ai_legal_html": ai_legal_html,
        "ai_risk_html": ai_risk_html,
        "compliance_score": int(latest_scan.report.score) if latest_scan and hasattr(latest_scan, 'report') else 100
    })


from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.views.decorators.http import require_POST
import secrets
from .models import Profile

@login_required
def profile_view(request):
    profile, created = Profile.objects.get_or_create(user=request.user)
    
    # Calculate stats
    scans = Scan.objects.filter(project__user=request.user)
    total_scans = scans.count()
    active_projects = Project.objects.filter(user=request.user).count()
    
    # Calculate average compliance score across all user's completed scans (using report score)
    completed_reports = [s.report for s in scans if hasattr(s, 'report') and s.report]
    if completed_reports:
        avg_score = int(sum(r.score for r in completed_reports) / len(completed_reports))
    else:
        avg_score = 100
        
    password_form = PasswordChangeForm(request.user)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'update_profile':
            username = request.POST.get('username')
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            email = request.POST.get('email')
            company = request.POST.get('company')
            default_wcag = request.POST.get('default_wcag_level', 'AA')
            
            # Simple validation & update
            if username:
                request.user.username = username
            request.user.first_name = first_name
            request.user.last_name = last_name
            request.user.email = email
            request.user.save()
            
            profile.company = company
            profile.default_wcag_level = default_wcag
            profile.save()
            
            messages.success(request, "Your profile settings have been successfully saved!")
            return redirect('profile')
            
        elif action == 'change_password':
            password_form = PasswordChangeForm(request.user, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, "Your password has been successfully updated!")
                return redirect('profile')
            else:
                messages.error(request, "Please correct the password validation errors below.")

    context = {
        'profile': profile,
        'total_scans': total_scans,
        'active_projects': active_projects,
        'avg_score': avg_score,
        'password_form': password_form,
    }
    return render(request, 'core/profile.html', context)


@login_required
@require_POST
def rotate_api_key_api(request, *args, **kwargs):
    profile, created = Profile.objects.get_or_create(user=request.user)
    new_token = f"wcag_live_{secrets.token_hex(24)}"
    profile.api_key = new_token
    profile.save()
    return JsonResponse({
        "success": True,
        "api_key": new_token
    })
